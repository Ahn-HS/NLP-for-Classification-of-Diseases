# -*- coding: utf-8 -*-

# 엑셀 파일 접근 패키지
import openpyxl

import time, re, collections, gensim
import operator
from collections import Counter
import math
import nltk
from nltk.collocations import *
from sklearn.feature_extraction.text import TfidfVectorizer

# 병렬 연산 처리
import multiprocessing
from krwordrank.word import KRWordRank
from gensim import corpora


def extraction_by_condition_value(variable):
    # final_pi_cc 엑셀 파일 로드
    exel_file = openpyxl.load_workbook('data/analysis/emr/EMR_RevisionPICC_First_Outpatient_revision.xlsx',
                                       data_only=True)
    excel_sheet = exel_file.get_sheet_by_name("CJ_TOTAL_MASKING")

    if variable == 'patient_yn':
        extraction(variable, '대상자에서 제외', excel_sheet)
        extraction(variable, '진단받은 환자군', excel_sheet)
        extraction(variable, '잠재적 환자군', excel_sheet)

    elif variable == 'icd10code':
        # extraction(variable, 'F32', excel_sheet)
        pass

    elif variable == 'present_age':
        pass

    elif variable == 'sex':
        extraction(variable, 'M', excel_sheet, 3, False)
        extraction(variable, 'F', excel_sheet, 3, False)

    elif variable == 'date':
        extraction(variable, '2007', excel_sheet, 2, False)
        extraction(variable, '2008', excel_sheet, 2, False)
        extraction(variable, '2009', excel_sheet, 2, False)
        extraction(variable, '2010', excel_sheet, 2, False)
        extraction(variable, '2011', excel_sheet, 2, False)
        extraction(variable, '2012', excel_sheet, 2, False)
        extraction(variable, '2013', excel_sheet, 2, False)
        extraction(variable, '2014', excel_sheet, 2, False)
        extraction(variable, '2015', excel_sheet, 2, False)
        extraction(variable, '2016', excel_sheet, 2, False)



''' 추출할 변수명, 해당 변수에 대한 조건명, 엑셀파일 객체, 해당 변수의 칼럼, 완전일치 또는 부분일치'''
def extraction(variable, condition_name, excel_sheet, column, like_computation):
    print("extraction by " + variable + " - " + condition_name)
    txtfile = open('data/analysis/extraction_by_condition/' + variable + "/" + condition_name +
                   '.txt', 'w', encoding='utf-8')

    cc_condition_by_text_list = []
    for r_index, row in enumerate(excel_sheet.rows):
        ''' condition( 조건 ) 에 따라서 PI와 CC 데이터 로드 '''
        if like_computation == True:
            if row[column].value == condition_name:
                if str(row[12].value) is not " ":
                    txtfile.writelines(' ' + str(row[12].value) + ' ')
                    txtfile.writelines("\n")
                    if row[32].value is not None:
                        cc_condition_by_text_list.append(' ' + str(row[32].value) + ' ')
        else:
            if condition_name in str(row[column].value):
                if row[12].value is not None:
                    txtfile.writelines(' ' + str(row[12].value) + ' ')
                    txtfile.writelines("\n")
                    if row[32].value is not None:
                        cc_condition_by_text_list.append(' ' + str(row[32].value) + ' ')

    for cc in cc_condition_by_text_list:
        txtfile.writelines(cc)
        txtfile.writelines("\n")

    txtfile.close()


def extraction_frequency(condition, filename):
    dic = {}
    with open('data/analysis/extraction_by_condition/' + condition + '/' + filename + '.txt', 'r',
              encoding='utf-8') as file:
        lines = file.readlines()
        for idx, line in enumerate(lines):
            words = line.split()
            for word in words:
                if word in dic:
                    dic[word] = dic[word] + 1
                else:
                    dic[word] = 1
    dic = {key: value for key, value in dic.items() if value > 2 and key.isnumeric() == False}

    sorted_freq_result = sorted(dic.items(), key=operator.itemgetter(1), reverse=True)
    return sorted_freq_result


def pos_tokenize_twitter():
    # 형태소 분석 토크나이즈
    try:
        ft = open('data/exeltotxt_senten_separ.txt', 'r', encoding='utf-8')
    except FileNotFoundError as e:
        print(str(e))

    spliter = Twitter()
    lines = ft.readlines()

    for line in lines:
        print(spliter.pos(line, norm=True, stem=True))


def word_modify(version):
    wb = openpyxl.load_workbook('data/ngram/excel/bi_tri_gram_result_' + version + '.xlsx', data_only=True)
    ws = wb.active

    r_ft = open('data/revision/revision_' + version + '.txt', 'r', encoding='utf-8')
    w_ft = open('data/revision/revision_' + str(int(version[0:1]) + 1) + version[1:] + '.txt', 'w', encoding='utf-8')
    w_ori_ft = open('data/revision/original/revision_' + str(int(version[0:1]) + 1) + version[1:] + '_ori.txt', 'w', encoding='utf-8')

    reviselist = []

    for lines in r_ft.readlines():
        reviselist.append(' ' + lines[:-1] + ' ')

    count = 1
    for row in ws.rows:
        if row[4].value is not None:
            front_w1 = row[0].value
            end_w2 = row[1].value
            replace_text = row[4].value
            txt_line_count = 0
            for list in reviselist:

                list = list.replace(' ' + str(front_w1) + ' ' + str(end_w2) + ' ', ' ' + replace_text + ' ')
                reviselist[txt_line_count] = list
                txt_line_count += 1
            print(str(count) + '. ' + str(front_w1) + ' ' + str(end_w2) + ' = ' + replace_text)
            count += 1

    for list in reviselist:
        w_ft.writelines(list.strip() + '\n')
        w_ori_ft.writelines(list.strip() + '\n')
    r_ft.close()
    w_ft.close()
    w_ori_ft.close()


def matching_with_EMR_MK():
    # 문장단위로 자른 utter_reponse 파일 로드
    exel_file = openpyxl.load_workbook('data/keyword_list/utter_response.xlsx', data_only=True)
    excel_sheet = exel_file.get_sheet_by_name("data")

    ################################ 해당 문장에 키워드가 존재하는지 마킹 ###################################
    for r_seq_1, row_1 in enumerate(excel_sheet.rows):
        if r_seq_1 > 0:
            print(str(row_1[4].value))

    ##################################### 카테고리 별 ####################################
    # for index, keyword in enumerate(klist):
    #     if keyword == '':
    #         count += 1
    #         row_count = 2
    #     ws.cell(row=3 + 5*(count) + (50*count), column=row_count, value=keyword)
    #     similar_list = model_similarity(file_name, keyword)
    #     if similar_list == 0:
    #         print('no similar word')
    #         continue
    #     for row, word in enumerate(similar_list):
    #         ws.cell(row=row + 5*(count+1) + (50*count), column=row_count, value=word[0])
    #         # 유사도추출
    #         # ws.cell(row=row + 5*(count+1) + (50*count), column=row_count+1, value=word[1])
    #     row_count += 2
    ######################################

    exel_file.save('data/EMR/marking_data.xlsx')


# 수정된 텍스트 파일에서 엑셀 파일로 변환
def txt_to_exel(load_version):
    exel_file = openpyxl.load_workbook('data/analysis/emr/EMR_RevisionPICC_First_Outpatient.xlsx', data_only=True)
    excel_sheet = exel_file.get_sheet_by_name("CJ_TOTAL_MASKING")

    with open('data/revision/revision_' + load_version + '.txt', 'r', encoding='utf-8') as file:
        lines = file.readlines()
        PI_or_CC = True
        idx_cc = 2
        for idx_pi, line in enumerate(lines):
            # print(line[:-1].strip())
            if "cc_cc_cc_cc_cc" in line:
                PI_or_CC = False
                # "cc_cc_cc_cc_cc" 부분 x
                continue
            if PI_or_CC:
                excel_sheet.cell(row=idx_pi + 2, column=13, value=line[:-1].strip())
            else:
                excel_sheet.cell(row=idx_cc, column=33, value=line[:-1].strip())
                idx_cc = idx_cc + 1

    exel_file.save('data/analysis/emr/EMR_RevisionPICC_First_Outpatient_revision.xlsx')


# n그램
def ngram(load_version):
    wb = openpyxl.Workbook()
    for idx, seq in enumerate(range(2, 3)):
        ngrams = word_ngram(seq, load_version)
        freqlist = make_freqlist(ngrams)
        freqlist = {key: value for key, value in freqlist.items() if value > 2}
        sorted_freqlist = sorted(freqlist.items(), key=operator.itemgetter(1), reverse=True)
        for idx, ngram in enumerate(sorted_freqlist):
            print(ngram)
            if idx > 100:
                break

        result_file = open('data/ngram/ngram_result/' + str(seq) + '_gram_result_' + load_version + '.txt', 'w',
                           encoding='utf-8')
        add_sheet = wb.create_sheet(str(seq), idx)

        for srt in range(len(sorted_freqlist)):
            result_file.writelines(str(sorted_freqlist[srt]) + "\r\n")
            split_list = str(sorted_freqlist[srt]).split("'")
            split_num = str(sorted_freqlist[srt]).split(",")
            separated, position = 1
            for n in range(0, seq):
                add_sheet.cell(row=srt + 1, column=position, value=split_list[separated])
                separated = separated + 2
                position = position + 1
            add_sheet.cell(row=srt + 1, column=position, value=split_num[seq][:-1])
        result_file.close()
    wb.save('data/analysis/ngram/n_gram_result_' + load_version + '.xlsx')


# n-gram 어절 분석
def word_ngram(num_gram, version):
    temp_text = open('data/revision/revision_temp.txt', 'r', encoding='utf-8')
    file = temp_text.read()

    sampling_file = ''.join(file.strip())
    sample_file = sampling_file.replace('\n', ' ').replace('\r', ' ')

    ###### txt 파일에서 영어와 숫자 제거 ######
    remove_num_doc = re.sub(r'\d', '', sample_file)
    remove_num_and_eng_doc = re.sub('[a-zA-Z]', '', remove_num_doc)
    # remove_num_and_eng_and_empty_doc = [x for x in remove_num_and_eng_doc if x]

    text = remove_num_and_eng_doc.split(' ')
    text = tuple([x for x in text if x])

    ngrams = [text[x:x + num_gram] for x in range(0, len(text))]

    return tuple(ngrams)


# n-gram 빈도 리스트
def make_freqlist(ngrams):
    freqlist = {}
    for ngram in ngrams:
        if (ngram in freqlist):
            freqlist[ngram] += 1
        else:
            freqlist[ngram] = 1
    return freqlist


def check_front_back_word(load_version, num_word, keywordlist):
    temp_file = open('data/revision/revision_temp.txt', 'r', encoding='utf-8')
    file_text = temp_file.read()

    sampling_file = ''.join(file_text)

    text_corpus = tuple(sampling_file.replace('\n', '').replace('\r', '').split(' '))
    for idx, word in enumerate(text_corpus):
        if word == "민망":
            sentence = ''
            for i in range(-num_word, num_word + 1):
                sentence += text_corpus[idx + i] + ' '
            sentence = sentence.replace('  ', ' ')
            print(sentence)
        # 해당 단어가 포함되어 있다면 앞뒤 n개 단어들이랑 리스트에 append


def listup_sentence_of_similar_word():
    pass


def get_PMI(load_version):
    '''
        'bigram_pmi': nltk.collocations.BigramAssocMeasures().pmi,
        'trigram_pmi': nltk.collocations.TrigramAssocMeasures().pmi,
    '''
    # trigram_measures = nltk.collocations.TrigramAssocMeasures()
    bigram_measures = nltk.collocations.BigramAssocMeasures()

    finder = BigramCollocationFinder.from_words(doc_split)
    # finder = TrigramCollocationFinder.from_words(word_tokenize(sampling_text))
    keywords = finder.score_ngrams(bigram_measures.pmi)
    # for idx, i in enumerate(finder.score_ngrams(bigram_measures.pmi)):


def get_likelihood_ratio(load_version):
    # trigram_measures = nltk.collocations.TrigramAssocMeasures()
    bigram_measures = nltk.collocations.BigramAssocMeasures()

    finder = BigramCollocationFinder.from_words(doc_split)
    keywords = finder.score_ngrams(bigram_measures.likelihood_ratio)

    # Group bigrams by first word in bigram.
    prefix_keys = collections.defaultdict(list)
    for key, scores in keywords:
        prefix_keys[key[0]].append((key[1], scores))

    for key in prefix_keys:
        prefix_keys[key].sort(key=lambda x: -x[1])

    word = "벤다졸"
    print(word, prefix_keys[word][:50])


def get_TFIDFbased_Concordance(load_version):
    corpus_list = []
    for i, line in enumerate(doc_split):
        if len(line.strip().split()) > 1:
            corpus_list.append(line.strip())

    vectorizer = TfidfVectorizer()
    # tfidf = TfidfVectorizer().fit_transform(corpus_list)
    # document_distances = (tfidf_matrix * tfidf_matrix.T)
    # print(document_distances.get_shape())


def get_concordance(load_version):
    ko = nltk.Text(emr_ko, name='data/revision/revision_' + load_version + '.txt')
    ko.vocab()


def kr_wordrank(load_version):
    # 1차원 리스트 구조 : 한 진료 데이터 단위
    file = open('data/revision/revision_' + load_version + '.txt', 'r', encoding='utf-8', newline='\n')
    list_corpus = []
    for sentence in file:
        list_corpus.append(sentence.strip())

    wordrank_extractor = KRWordRank(
        min_count=5,  # 단어의 최소 출현 빈도수 (그래프 생성 시)
        max_length=10,  # 단어의 최대 길이
        verbose=True
    )
    beta = 0.85  # PageRank의 decaying factor beta
    max_iter = 10
    keywords, rank, graph = wordrank_extractor.extract(list_corpus, beta, max_iter)

    for word, r in sorted(keywords.items(), key=lambda x: x[1], reverse=True)[:100]:
        print('%8s:\t%.4f' % (word, r))


def get_texts_scores(fname):
    with open(fname, encoding='utf-8') as f:
        docs = [doc.lower().replace('\n', '').split('\t') for doc in f]
        docs = [doc for doc in docs if len(doc) == 2]

        if not docs:
            return [], []

        texts, scores = zip(*docs)
        return list(texts), list(scores)


def skipgrams(load_version):
    file = open('data/revision/revision_' + load_version + '.txt', 'r', encoding='utf-8', newline='\n')
    file_text = file.read()
    remove_num_doc = re.sub(r'\d', '', file_text)
    remove_eng_doc = re.sub('[a-zA-Z]', '', remove_num_doc)
    doc_split = remove_eng_doc.split()

    back_window = 2
    front_window = 2
    skipgram_counts = Counter()

    for idx, word in enumerate(doc_split):
        icw_min = max(0, idx - back_window)
        icw_max = min(len(doc_split) - 1, idx + front_window)
        icws = [ii for ii in range(icw_min, icw_max + 1) if ii != idx]
        for icw in icws:
            skipgram = (doc_split[idx], doc_split[icw])
            skipgram_counts[skipgram] += 1
    print('number of skipgrams: {}'.format(len(skipgram_counts)))
    # print('most common: {}'.format(skipgram_counts.most_common(100)))
    for word in skipgram_counts.most_common(100):
        print(word)


def LDA_modeling(load_version, topic_num):
    with open('data/revision/revision_' + load_version + '.txt', 'r', encoding='utf-8', newline='\n') as f:
        lines = f.readlines()

    corpus_list = []
    for idx, line in enumerate(lines):
        remove_num_line = re.sub(r'\d', '', line)
        remove_num_n_eng_line = re.sub('[a-zA-Z]', '', remove_num_line)

        corpus_list.append(remove_num_n_eng_line.split())

    dictionary = corpora.Dictionary(corpus_list)
    corpus = [dictionary.doc2bow(text) for text in corpus_list]
    ldamodel = gensim.models.ldamodel.LdaModel(corpus, num_topics=topic_num, id2word=dictionary, passes=20)
    ldamodel.save('data/analysis/LDAmodel/' + load_version + "_" + str(topic_num))


def get_LDA(load_version, topic_num):
    ldamodel = gensim.models.ldamodel.LdaModel.load('data/analysis/LDAmodel/' + load_version + "_" + str(topic_num))
    # ldamodel.update(파일이름)
    for i in range(0, topic_num):
        print(ldamodel.show_topic(i, 20))
    # print(ldamodel.print_topics(num_topics=10, num_words=10))


if __name__ == "__main__":
    start_time = time.time()
    load_save_txt_version = '48t'

    ############################# 조건 추출 ##################################

    ''' 특정 셀이 특정 조건에 부합한다면 셀 값을 로드 '''
    # 진단 받은 별로 구분
    # extraction_by_condition_value('patient_yn')
    # icd10 코드 별로 구분
    # extraction_by_condition_value('icd10code')
    # 나이별로 구분
    # extraction_by_condition_value('present_age')
    # 성별로 구분
    # extraction_by_condition_value('sex')
    # 연도별로 구분
    # extraction_by_condition_value('date')
    ######################################################

    ''' pi_cc_final 데이터 병합 '''
    # pi_cc_combine()

    ''' EMR 데이터에서 증상 키워드와 매칭되는 문장 검색 '''
    # matching_with_EMR_MK()

    ''' n 그램 돌리고 n 그램 결과 (텍스트 파일)과 (엑셀 파일)로 저장 '''
    ''' 영어와 숫자 제거 버전 '''
    # ngram(load_save_txt_version)

    ''' (리스트업)n 그램 결과 엑셀 파일에서 전문의 키워드가 속한 n 그램만 리스트업 '''
    # listup_ngram_keyword_extraction(load_save_txt_version, depression_keyword_list)

    ''' revision.txt 파일에서 특정 단어에 대해 주변 단어들 추출 및 저장( 추출할 주변 단어 개수, 키워드 리스트 ) '''
    # check_front_back_word(load_save_txt_version, 3, depression_keyword_list)

    ''' 빈도수를 체크하고 txt 파일로 저장 True: 엑셀파일에서, False: 텍스트파일에서'''
    # word_frequency(load_save_txt_version, False)

    ''' similar_result 엑셀 파일에서 추출된 유사어를 기준으로 주변 단어들 리스트 업'''
    # listup_sentence_of_similar_word(load_save_txt_version)

    ''' nltk 패키지를 사용하여 PMI 값을 계산한다. '''
    # get_PMI(load_save_txt_version)
    # get_PMI2(load_save_txt_version)

    ''' nltk 패키지를 사용하여 likelihood ratio 계산 '''
    # get_likelihood_ratio(load_save_txt_version)

    ''' sklearn 패키지를 사용하여 TF-IDF를 계산한다 '''
    # get_TFIDF(load_save_txt_version)

    '''대용량 csv 파일 처리 25,545,842 건 / 1분 38초 '''
    # open_csv()

    ''' 한 문장에서 동시 출현에 정보 추출 '''
    # get_concordance(load_save_txt_version)

    ''' kr-wordrank 구현 '''
    # kr_wordrank(load_save_txt_version)

    ''' skipgram '''
    # skipgrams(load_save_txt_version)

    ''' LDA 모델링 '''
    # LDA_modeling(load_save_txt_version, 20)
    # get_LDA(load_save_txt_version, 20)
    # visualize_LDA(load_save_txt_version)

    ''' LSA 모델링 '''
    # LSA_modeling(load_save_txt_version)
    # get_LSA(load_save_txt_version)

    ''' 불용어 제거 '''
    # remove_stopword()

    e = int(time.time() - start_time)
    print('\r\n{:02d}:{:02d}:{:02d}'.format(e // 3600, (e % 3600 // 60), e % 60))
